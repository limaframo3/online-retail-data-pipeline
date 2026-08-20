"""Append controlled product changes to the raw Excel source for an SCD2 demo.

Expected demonstration flow:

    python run_pipeline.py
    python scripts/apply_scd2_demo_changes.py
    python run_pipeline.py

The first pipeline run creates the initial warehouse. This script then appends
two new, valid sales rows whose product descriptions differ from the current
versions in ``dim_product``. On the second pipeline run, those rows become new
SCD Type 2 product versions.

The source workbook is backed up before it is changed. Use ``--restore`` to
return it to its original state after the demonstration.
"""

from __future__ import annotations

import argparse
import json
import logging
import os
import shutil
import sys
from datetime import datetime, time, timedelta
from pathlib import Path
from typing import Any


# =========================================================
# CONFIGURATION
# =========================================================
BASE_DIR = Path(__file__).resolve().parent.parent
CONFIG_PATH = BASE_DIR / "config.json"
DEFAULT_EXCEL_PATH = BASE_DIR / "data" / "raw" / "Online Retail.xlsx"
DEFAULT_DW_PATH = BASE_DIR / "db" / "DW_Online_Retail.db"
LOG_PATH = BASE_DIR / "logs" / "scd2_demo.log"
BACKUP_SUFFIX = ".before_scd2_demo.xlsx"
DEMO_INVOICE_PREFIX = "SCD2DEMO"
DEMO_DESCRIPTION_SUFFIX = " [SCD2 DEMO UPDATE]"
DEMO_PRODUCT_COUNT = 2
PREFERRED_STOCKCODES = ("85123A", "22423")

REQUIRED_COLUMNS = {
    "invoiceno",
    "stockcode",
    "description",
    "quantity",
    "invoicedate",
    "unitprice",
    "customerid",
    "country",
}


# =========================================================
# LOGGER
# =========================================================
def setup_logger() -> None:
    """Configure a dedicated log for the SCD2 demo utility."""
    LOG_PATH.parent.mkdir(parents=True, exist_ok=True)

    logging.basicConfig(
        filename=str(LOG_PATH),
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )


def parse_args() -> argparse.Namespace:
    """Parse command-line options."""
    parser = argparse.ArgumentParser(
        description=(
            "Append controlled product-description changes to the raw "
            "Online Retail workbook for an SCD Type 2 demonstration."
        )
    )
    parser.add_argument(
        "--restore",
        action="store_true",
        help="Restore the raw workbook from the backup created by this script.",
    )
    return parser.parse_args()


def load_config() -> dict[str, Any]:
    """Load config.json when it is available."""
    if not CONFIG_PATH.exists():
        return {}

    with CONFIG_PATH.open("r", encoding="utf-8") as config_file:
        return json.load(config_file)


def resolve_project_path(config_value: str | None, default: Path) -> Path:
    """Resolve a configured path relative to the project root."""
    if not config_value:
        return default

    configured_path = Path(config_value)
    if configured_path.is_absolute():
        return configured_path
    return BASE_DIR / configured_path


def get_paths() -> tuple[Path, Path, Path]:
    """Return the raw Excel, warehouse, and backup paths."""
    config = load_config()
    configured_paths = config.get("paths", {})

    excel_path = resolve_project_path(
        configured_paths.get("excel"),
        DEFAULT_EXCEL_PATH,
    )
    dw_path = resolve_project_path(
        configured_paths.get("database"),
        DEFAULT_DW_PATH,
    )
    backup_path = excel_path.with_name(
        f"{excel_path.stem}{BACKUP_SUFFIX}"
    )
    return excel_path, dw_path, backup_path


def normalize_header(value: Any) -> str:
    """Normalize an Excel column name for reliable matching."""
    return "".join(
        character
        for character in str(value or "").strip().lower()
        if character.isalnum()
    )


def normalize_stockcode(value: Any) -> str:
    """Normalize stock codes read from Excel or DuckDB for matching."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def validate_initial_warehouse(dw_path: Path) -> dict[str, str]:
    """Return current product descriptions from a completed initial load."""
    import duckdb

    if not dw_path.exists():
        raise FileNotFoundError(
            f"Data Warehouse not found: {dw_path}\n"
            "Run 'python run_pipeline.py' once before applying demo changes."
        )

    connection = duckdb.connect(str(dw_path), read_only=True)
    try:
        columns = {
            row[1]
            for row in connection.execute(
                "PRAGMA table_info('dim_product');"
            ).fetchall()
        }
        missing = {
            "stockcode",
            "description",
            "effective_from",
            "effective_to",
            "is_current",
        } - columns
        if missing:
            missing_text = ", ".join(sorted(missing))
            raise RuntimeError(
                "dim_product is not using the expected SCD Type 2 schema. "
                f"Missing columns: {missing_text}."
            )

        rows = connection.execute("""
            SELECT
                CAST(stockcode AS VARCHAR),
                CAST(description AS VARCHAR)
            FROM dim_product
            WHERE is_current = TRUE
            ORDER BY stockcode;
        """).fetchall()
    finally:
        connection.close()

    if not rows:
        raise RuntimeError(
            "dim_product contains no current products. "
            "Complete the initial pipeline run first."
        )

    return {
        normalize_stockcode(stockcode): str(description).strip()
        for stockcode, description in rows
    }


def as_datetime(value: Any) -> datetime | None:
    """Convert supported Excel date values to datetime."""
    if isinstance(value, datetime):
        return value
    if value is None:
        return None

    text = str(value).strip()
    for date_format in (
        "%Y-%m-%d %H:%M:%S",
        "%m/%d/%Y %H:%M",
        "%d/%m/%Y %H:%M",
    ):
        try:
            return datetime.strptime(text, date_format)
        except ValueError:
            continue
    return None


def as_positive_number(value: Any) -> float | None:
    """Return a positive numeric value, or None when invalid."""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def select_demo_products(
    candidates: dict[str, dict[str, Any]],
    current_product_count: int,
    matching_stockcode_count: int,
) -> list[dict[str, Any]]:
    """Choose preferred products first, then deterministic fallbacks."""
    selected: list[dict[str, Any]] = []

    for stockcode in PREFERRED_STOCKCODES:
        if stockcode in candidates:
            selected.append(candidates[stockcode])

    if len(selected) < DEMO_PRODUCT_COUNT:
        already_selected = {
            str(item["stockcode"])
            for item in selected
        }
        fallbacks = sorted(
            (
                item
                for stockcode, item in candidates.items()
                if stockcode not in already_selected
            ),
            key=lambda item: str(item["stockcode"]),
        )
        selected.extend(
            fallbacks[: DEMO_PRODUCT_COUNT - len(selected)]
        )

    if len(selected) < DEMO_PRODUCT_COUNT:
        raise RuntimeError(
            f"Only {len(selected)} suitable product(s) were found; "
            f"{DEMO_PRODUCT_COUNT} are required for the demo. "
            f"Current products in dim_product: {current_product_count}. "
            "Products found in both the workbook and dim_product: "
            f"{matching_stockcode_count}. Verify that the pipeline and this "
            "script use the same source workbook."
        )

    return selected


def append_demo_changes(
    excel_path: Path,
    backup_path: Path,
    current_products: dict[str, str],
) -> list[dict[str, Any]]:
    """Back up the workbook and append controlled SCD2 change rows."""
    from openpyxl import load_workbook

    if not excel_path.exists():
        raise FileNotFoundError(f"Raw Excel file not found: {excel_path}")

    workbook = load_workbook(excel_path)
    worksheet = workbook.active

    headers = {
        normalize_header(cell.value): cell.column
        for cell in worksheet[1]
        if cell.value is not None
    }
    missing_columns = REQUIRED_COLUMNS - set(headers)
    if missing_columns:
        workbook.close()
        missing_text = ", ".join(sorted(missing_columns))
        raise RuntimeError(
            f"The raw workbook is missing required columns: {missing_text}."
        )

    candidates: dict[str, dict[str, Any]] = {}
    maximum_invoice_date: datetime | None = None
    matching_stockcodes: set[str] = set()

    for row_number in range(2, worksheet.max_row + 1):
        invoice_number = worksheet.cell(
            row=row_number,
            column=headers["invoiceno"],
        ).value
        if str(invoice_number or "").startswith(DEMO_INVOICE_PREFIX):
            workbook.close()
            raise RuntimeError(
                "SCD2 demo rows already exist in the raw workbook. "
                "Run the pipeline to load them, or use '--restore' first."
            )

        invoice_date = as_datetime(
            worksheet.cell(
                row=row_number,
                column=headers["invoicedate"],
            ).value
        )
        if invoice_date is None:
            continue

        if maximum_invoice_date is None or invoice_date > maximum_invoice_date:
            maximum_invoice_date = invoice_date

        stockcode = normalize_stockcode(
            worksheet.cell(
                row=row_number,
                column=headers["stockcode"],
            ).value
        )
        if stockcode not in current_products:
            continue

        matching_stockcodes.add(stockcode)

        # Use the description from the current DW version. The raw row is
        # needed only to obtain valid transaction attributes. Requiring its
        # historical description to match the current version exactly can
        # incorrectly discard otherwise suitable products.
        current_description = current_products[stockcode]

        quantity = as_positive_number(
            worksheet.cell(
                row=row_number,
                column=headers["quantity"],
            ).value
        )
        unit_price = as_positive_number(
            worksheet.cell(
                row=row_number,
                column=headers["unitprice"],
            ).value
        )
        customer_id = worksheet.cell(
            row=row_number,
            column=headers["customerid"],
        ).value
        country = worksheet.cell(
            row=row_number,
            column=headers["country"],
        ).value

        if (
            quantity is None
            or unit_price is None
            or customer_id in (None, "")
            or country in (None, "")
        ):
            continue

        previous_candidate = candidates.get(stockcode)
        if (
            previous_candidate is None
            or invoice_date > previous_candidate["invoicedate"]
        ):
            candidates[stockcode] = {
                "stockcode": stockcode,
                "description": current_description,
                "unitprice": unit_price,
                "customerid": customer_id,
                "country": country,
                "invoicedate": invoice_date,
            }

    if maximum_invoice_date is None:
        workbook.close()
        raise RuntimeError("No valid invoice dates were found in the workbook.")

    selected = select_demo_products(
        candidates=candidates,
        current_product_count=len(current_products),
        matching_stockcode_count=len(matching_stockcodes),
    )
    demo_date = datetime.combine(
        maximum_invoice_date.date() + timedelta(days=1),
        time(hour=9),
    )
    demo_invoice = f"{DEMO_INVOICE_PREFIX}-{demo_date:%Y%m%d}"

    if not backup_path.exists():
        shutil.copy2(excel_path, backup_path)

    changes: list[dict[str, Any]] = []
    for offset, product in enumerate(selected):
        new_description = (
            product["description"] + DEMO_DESCRIPTION_SUFFIX
        )
        new_row: list[Any] = [None] * worksheet.max_column

        values = {
            "invoiceno": demo_invoice,
            "stockcode": product["stockcode"],
            "description": new_description,
            "quantity": 1,
            "invoicedate": demo_date + timedelta(minutes=offset),
            "unitprice": product["unitprice"],
            "customerid": product["customerid"],
            "country": product["country"],
        }
        for column_name, value in values.items():
            new_row[headers[column_name] - 1] = value

        worksheet.append(new_row)
        changes.append(
            {
                "stockcode": product["stockcode"],
                "old_description": product["description"],
                "new_description": new_description,
                "effective_from": values["invoicedate"],
                "invoice": demo_invoice,
            }
        )

    temporary_path = excel_path.with_name(
        f".{excel_path.stem}.scd2_demo.tmp{excel_path.suffix}"
    )
    try:
        workbook.save(temporary_path)
        workbook.close()
        os.replace(temporary_path, excel_path)
    except Exception:
        workbook.close()
        temporary_path.unlink(missing_ok=True)
        raise

    return changes


def restore_source(excel_path: Path, backup_path: Path) -> None:
    """Restore the original source workbook from the demo backup."""
    if not backup_path.exists():
        raise FileNotFoundError(
            f"Demo backup not found: {backup_path}"
        )
    shutil.copy2(backup_path, excel_path)
    logging.info(
        "SCD2 demo source restored. Backup: %s, destination: %s",
        backup_path,
        excel_path,
    )


def main() -> None:
    """Apply or restore the controlled SCD Type 2 demonstration."""
    setup_logger()
    args = parse_args()
    excel_path, dw_path, backup_path = get_paths()

    logging.info("=" * 78)
    logging.info("Starting SCD Type 2 demo utility")
    logging.info("Source workbook: %s", excel_path)
    logging.info("Data Warehouse: %s", dw_path)

    if args.restore:
        logging.info("Restore mode requested")
        restore_source(excel_path, backup_path)
        print("\nSCD TYPE 2 DEMO SOURCE RESTORED")
        print(f"Workbook: {excel_path}")
        print(f"Log file: {LOG_PATH}")
        print("Run 'python run_pipeline.py' to rebuild from the original source.")
        logging.info("SCD Type 2 demo restore completed successfully")
        return

    logging.info("Validating initial Data Warehouse")
    current_products = validate_initial_warehouse(dw_path)
    logging.info(
        "Current products available in dim_product: %s",
        len(current_products),
    )

    changes = append_demo_changes(
        excel_path=excel_path,
        backup_path=backup_path,
        current_products=current_products,
    )

    logging.info("Demo source backup available at: %s", backup_path)
    for change in changes:
        logging.info(
            "SCD2 demo change applied. Stockcode: %s, "
            "old description: %s, new description: %s, "
            "effective from: %s, invoice: %s",
            change["stockcode"],
            change["old_description"],
            change["new_description"],
            change["effective_from"],
            change["invoice"],
        )

    logging.info(
        "SCD Type 2 demo changes completed successfully. "
        "Products changed: %s",
        len(changes),
    )

    print("\n" + "=" * 78)
    print("SCD TYPE 2 DEMO CHANGES APPLIED")
    print("=" * 78)
    for change in changes:
        print(f"Stock code:      {change['stockcode']}")
        print(f"Old description: {change['old_description']}")
        print(f"New description: {change['new_description']}")
        print(f"Effective from:  {change['effective_from']}")
        print(f"Demo invoice:    {change['invoice']}")
        print("-" * 78)

    print(f"Backup: {backup_path}")
    print(f"Log file: {LOG_PATH}")
    print("Next step: python run_pipeline.py")
    print("The second run should close each previous product version and create")
    print("a new current version in dim_product.")
    print("=" * 78)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        logging.exception("Unable to apply SCD2 demo changes: %s", error)
        print(f"\nUnable to apply SCD2 demo changes: {error}", file=sys.stderr)
        print(f"Review the log file: {LOG_PATH}", file=sys.stderr)
        sys.exit(1)